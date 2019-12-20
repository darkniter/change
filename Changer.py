import openpyxl
import csv
import config
from profilehooks import timecall, profile
# from geolocation import geoadressation
from os import remove
from os import path
from err_decorator import error_decorator


err_rec = []
# geodata = []

result = [['P_STREET', 'P_HOUSE', 'P_MODEL', 'P_IP_OLD', 'P_IP',
            'P_VECTOR', 'P_UPLINK', 'P_MAC', 'P_VLAN', 'P_DATE_SETUP',
            'P_DATE_INSTALL', 'P_FLATS', 'P_DOOR', 'P_FLOOR',
            'P_DESCRIPTION', 'P_RESERVED1', 'P_RESERVED2', 'P_RESERVED3',
            'P_TRANSIT', 'P_REMOVED', 'P_HOUSE_ID']]
double = []


# функция для создания и заполнения массива записей домов
# (ввод: имя внутри функции, вывод: маcсив записей)
@timecall
def houses_init():
    houses = []
    houses_wb = openpyxl.load_workbook(config.HOUSES)
    houses_sn = houses_wb.sheetnames[0]
    w_sheet = houses_wb[houses_sn]
    houses_list = w_sheet.rows
    next(houses_list)
    for row in houses_list:
        cols = []
        for init in row:
            cols.append(init.value)
        houses.append(cols)
    return(houses)


def houses_filter(town, houses):
    houses_filtred = []
    for row in houses:
        if (row[1] in town):
            houses_filtred.append(row)
    return houses_filtred


@timecall
def hardware_init(fname, sheet):

    hardware = []

    hardware_wb = openpyxl.load_workbook(fname)
    hardware_list = hardware_wb[sheet].rows

    next(hardware_list)

    for row in hardware_list:
        cols = [None]*20
        for init in row:
            cols[init.column-1] = init.value

            if (((init.column == 10) or (init.column == 11)) and
               (init.value is not None)):
                cols[init.column-1] = str(init.value).split(' ')[0]
            if (((init.column == 1) or (init.column == 2)) and (init.font.strike is True)):
                cols[19] = 1
            if (init.column == 4 or init.column == 5):
                if(init.fill.fgColor.rgb == 'FF00B0F0'):
                    cols[18] = 1
        if not (cols[0] is None):
            hardware.append(cols)
    return(hardware)


@timecall
@error_decorator()
def result_init(town, fname, sheet, houses):
    houses_town = houses_filter(town, houses)
    _err = []
    _double = []
    _result = []

    hardware = hardware_init(fname, sheet)
    print("hardware:", len(hardware))
    for init in hardware:
        res_tmp = []
        try:
            number_hard = ((str(init[1]).split(','))[0].split('.'))[0]
        except BaseException:
            pass

        street_hard = init[0]

        for row in houses_town:

            try:
                number_house_arr = str(row[3]).split()
                row[3] = number_house_arr[0]
                street_house = row[2]
            except BaseException:
                print('err try')

            try:
                if(
                    (number_house_arr[1][0].isalpha()) or
                    ('/' in number_house_arr[1])
                ):
                    row[3] = "".join(number_house_arr)

            except BaseException:
                pass

            number_house = row[3]

            street_hard_tmp = street_hard.upper().strip()

            if ((street_house == 'УЛ. .') or (street_house == 'ул. .') or
               (street_house == 'Ул. .')):
                street_house = row[1]

            if len(street_hard_tmp.split('.')) < 2:
                street_hard_tmp = 'УЛ. ' + street_hard_tmp

            street_house_tmp = street_house.upper().strip()

            if (
                (street_house_tmp == street_hard_tmp) &
                (number_hard.upper().strip() == number_house.upper().strip()) &
                (row[1] in town)
                                ):

                res_tmp.append([init[0]]+[str(init[1])]+init[2:]+[int(row[0])])
                break

        if not res_tmp:
            # print(init[:4])
            _err += [init]
        else:
            if (len(res_tmp) > 1):
                for record in res_tmp:
                    if (record[18] or record[19]) == '1':
                        _double += res_tmp
                    else:
                        _result += res_tmp
                # print(type(res_tmp[1]))
            else:
                _result += res_tmp

            # adress = (res_tmp[0][5] + ', ' + res_tmp[0][1] + ', ' +
            #           res_tmp[0][2])

            # location = geoadressation(adress)
            # geodata.append([str(res_tmp[0][4])] + [location])

    print("result:", len(_result))
    return _result, _err, _double


@timecall
def out_file(result, namefile):
    # try:
    #     with open(namefile + '.csv', newline='') as newfile:
    #         scvwr = csv.writer(newfile, delimiter=';')
    #         for row in result:
    #             scvwr.writerow(row)
    # except BaseException:

    if path.isfile(config.DIR + namefile + '.csv'):
        remove(config.DIR + namefile + '.csv')

    with open(

                config.DIR + namefile + '.csv',
                'a+',
                newline=''

            ) as newfile:

        scvwr = csv.writer(newfile, delimiter=';', quoting=csv.QUOTE_MINIMAL)
        scvwr.writerows(result)
    print(namefile + ' done')
    newfile.close()

@profile
def main():


    houses = houses_init()

    result_init(
        'Г. ЛИКИНО-ДУЛЕВО',
        config.HARDWARE,
        'Комутаторы ЛД', houses
        )

    result_init(
        'Г. ОРЕХОВО-ЗУЕВО',
        config.HARDWARE,
        'Комутаторы ОЗ', houses
        )


    out_file(result, 'result_OZ')
    out_file(err_rec, 'Err_OZ')
    out_file(double, 'Дубли')

    result_init('Г. КУРОВСКОЕ',
                                         config.HARDWARE,
                                         'Комутаторы КУ', houses)



    result_init(
        'Д. КАБАНОВО',
        config.HARDWARE,
        'Комутаторы КБ', houses
        )



    result_init(
        ['Д. ДЕМИХОВО', 'Д. НАЖИЦЫ', 'Д. КРАСНАЯ ДУБРАВА'],
        config.HARDWARE,
        'Комутаторы ДМ', houses)

    # # global err_rec
    # err_rec = []
    # # geodata = []
    # # global result
    # result = [['P_STREET', 'P_HOUSE', 'P_MODEL', 'P_IP_OLD', 'P_IP',
    #             'P_VECTOR', 'P_UPLINK', 'P_MAC', 'P_VLAN', 'P_DATE_SETUP',
    #             'P_DATE_INSTALL', 'P_FLATS', 'P_DOOR', 'P_FLOOR',
    #             'P_DESCRIPTION', 'P_RESERVED1', 'P_RESERVED2', 'P_RESERVED3',
    #             'P_TRANSIT', 'P_REMOVED', 'P_HOUSE_ID']]
    # # global double
    # double = []



    result_init(
        'Г. ОРЕХОВО-ЗУЕВО',
        config.HARDWARE,
        'Broken ОЗ', houses
        )

    out_file(result, 'result_Broken')
    out_file(err_rec, 'Err_Broken')
    out_file(double, 'Дубли')


if __name__ == "__main__":
    main()
